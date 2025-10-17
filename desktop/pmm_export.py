"""Export PMM Workloads into an Excel template.

This module copies or loads the provided template (pmm_sample.xlsx), sets B2 to the
earliest week date, and writes man_week values into the matrix where rows are work
categories (column A) and columns step by 7 days from B2 across ~3 years.

Requirements: openpyxl
"""

from __future__ import annotations

from typing import Any, Dict, Iterable, Tuple, List
import datetime as dt
import os
import csv


def _to_date(iso: str) -> dt.date:
    return dt.date.fromisoformat(iso)


def _week_diff(base: dt.date, target: dt.date) -> int:
    delta = target - base
    # Expect Monday-based iso dates; enforce 7-day steps
    return delta.days // 7


def _to_monday(d: dt.date) -> dt.date:
    return d - dt.timedelta(days=d.weekday())


def _aggregate(records: Iterable[dict], phases: Iterable[dict] | None = None) -> Tuple[Dict[Tuple[str, str], float], dt.date]:
    """Aggregate man_week by (category_name, week) and return earliest week (Monday).

    earliest is the minimum of:
      - PMM workload weeks (already Mondays)
      - Phase end_date normalized to Monday

    Returns:
        (agg, earliest_date)
    """
    earliest: dt.date | None = None
    agg: Dict[Tuple[str, str], float] = {}
    for r in records:
        week = r.get("week")
        if not week:
            continue
        week_date = _to_date(week)
        if earliest is None or week_date < earliest:
            earliest = week_date
        wc = r.get("work_category") or {}
        name = wc.get("name") or "Unassigned"
        val = r.get("man_week") or 0
        try:
            fval = float(val)
        except Exception:
            fval = 0.0
        key = (name, week)
        agg[key] = agg.get(key, 0.0) + fval

    # Consider phases' end_date as well
    if phases:
        for p in phases:
            end_iso = p.get("end_date")
            if not end_iso:
                continue
            try:
                end_date = dt.date.fromisoformat(end_iso)
            except Exception:
                continue
            end_monday = _to_monday(end_date)
            if earliest is None or end_monday < earliest:
                earliest = end_monday
    if earliest is None:
        # Default to today Monday if no records
        today = dt.date.today()
        earliest = today - dt.timedelta(days=today.weekday())
    return agg, earliest


def export_pmm_workloads_to_xlsx(
    payload: dict,
    template_path: str,
    save_path: str,
) -> Dict[str, Any]:
    """Write PMM workloads into an Excel file based on a template.

    Args:
        payload: { subproject: {...}, records: IPMMWorkload[] }
        template_path: path to pmm_sample.xlsx
        save_path: where to save the populated xlsx
    """
    try:
        try:
            import openpyxl  # type: ignore
        except Exception as e:
            return {"success": False, "error": f"openpyxl is required: {e}"}

        if not os.path.exists(template_path):
            return {"success": False, "error": f"Template not found: {template_path}"}

        records = (payload or {}).get("records") or []
        phases: List[dict] = (payload or {}).get("phases") or []
        agg, earliest = _aggregate(records, phases)

        # Load template and save to new path after modifications
        wb = openpyxl.load_workbook(template_path)
        # Prefer an existing sheet named "Manpower Sheet"; otherwise rename active
        desired_sheet_name = "Manpower Sheet"
        if desired_sheet_name in wb.sheetnames:
            ws = wb[desired_sheet_name]
        else:
            ws = wb.active
            try:
                ws.title = desired_sheet_name
            except Exception:
                # If renaming fails, continue using active sheet
                pass

        # Set B2 to earliest date (as a date cell). Excel will propagate via formulas.
        ws.cell(row=2, column=2).value = earliest

        # Build a lookup for category names in column A (search a reasonable range)
        name_to_row: Dict[str, int] = {}
        # Assume categories start at row 3; scan first 1000 rows
        for r in range(1, 2000):
            v = ws.cell(row=r, column=1).value
            if isinstance(v, str) and v.strip():
                name_to_row[v.strip()] = r

        # Write aggregated values
        for (cat_name, week_iso), value in agg.items():
            row = name_to_row.get(cat_name)
            if row is None:
                # If category not present in template, skip (or append). For now, skip.
                # Could be enhanced to append at the end: row = max(name_to_row.values(), default=2) + 1
                continue
            week_date = _to_date(week_iso)
            col_offset = _week_diff(earliest, week_date)
            col = 2 + col_offset  # B=2
            # Write number into the target cell
            ws.cell(row=row, column=col).value = float(value)

        # Write phases into rows 3 (milestone=True) and 4 (milestone=False)
        if phases:
            for p in phases:
                name = p.get("name")
                end_iso = p.get("end_date")
                if not name or not end_iso:
                    continue
                try:
                    end_date = dt.date.fromisoformat(end_iso)
                except Exception:
                    continue
                col_offset = _week_diff(earliest, _to_monday(end_date))
                col = 2 + col_offset  # B=2
                row = 3 if bool(p.get("milestone")) else 4
                ws.cell(row=row, column=col).value = str(name)

        
        # Create or reuse a sheet named "4QV2" to tabulate raw records
        try:
            ws_list_name = "4QV2"
            if ws_list_name in wb.sheetnames:
                ws_list = wb[ws_list_name]
            else:
                # To preserve formatting, require a preformatted sheet in the template
                raise RuntimeError("Template sheet '4QV2' not found. Please add it to the template to preserve formatting.")

            # Capture prototype styles from header or first data row to replicate formatting
            proto_row_idx = 2 if ws_list.max_row and ws_list.max_row >= 2 else 1
            proto_styles: List[Tuple[Any, Any, Any, Any, Any, Any]] = []
            for c_idx in range(1, 8):
                cell = ws_list.cell(row=proto_row_idx, column=c_idx)
                proto_styles.append((cell.font, cell.fill, cell.border, cell.alignment, cell.number_format, cell.protection))

            # Clear only values in data rows (keep styles)
            if ws_list.max_row and ws_list.max_row >= 2:
                for r_idx in range(2, ws_list.max_row + 1):
                    for c_idx in range(1, 8):  # A..G
                        ws_list.cell(row=r_idx, column=c_idx).value = None

            # Pre-compute phase windows based on end_date order
            phase_windows: List[Tuple[dt.date, dt.date, str]] = []
            try:
                # collect phases with valid end_date
                _phs = []
                for p in phases or []:
                    # Exclude milestones from phase windows
                    if bool(p.get("milestone")):
                        continue
                    end_iso = p.get("end_date")
                    name = p.get("name")
                    if not end_iso or not name:
                        continue
                    try:
                        end_dt = dt.date.fromisoformat(end_iso)
                    except Exception:
                        continue
                    _phs.append((end_dt, str(name)))
                # sort by end_date asc
                _phs.sort(key=lambda x: x[0])
                # build windows [prev_end+1, end]
                prev_end: dt.date | None = None
                for end_dt, nm in _phs:
                    start_dt = (prev_end + dt.timedelta(days=1)) if prev_end else dt.date.min
                    phase_windows.append((start_dt, end_dt, nm))
                    prev_end = end_dt
            except Exception:
                phase_windows = []

            # Sort records by work_category using Manpower Sheet order (name_to_row), then by week (ascending)
            def _rec_wc_name(rec: dict) -> str:
                wc = rec.get("work_category")
                if wc is None:
                    wc = rec.get("work_cotegory")
                if isinstance(wc, dict):
                    return wc.get("name") or "Unassigned"
                return wc or "Unassigned"

            def _rec_week_date(rec: dict) -> dt.date:
                w = rec.get("week")
                if not w:
                    return dt.date.max
                try:
                    return _to_date(w)
                except Exception:
                    return dt.date.max

            def _rec_wc_row(rec: dict) -> int:
                nm = _rec_wc_name(rec)
                # Place unknown categories at the end
                return name_to_row.get(nm, 10**9)

            sorted_records = sorted(records, key=lambda rec: (_rec_wc_row(rec), _rec_week_date(rec), _rec_wc_name(rec)))

            row_idx = 2
            for r in sorted_records:
                name = r.get("name")
                wc = r.get("work_category")
                # Fallback for potential key typo
                if wc is None:
                    wc = r.get("work_cotegory")
                if isinstance(wc, dict):
                    wc_name = wc.get("name")
                else:
                    wc_name = wc
                try:
                    mw = float(r.get("man_week") or 0)
                except Exception:
                    mw = 0.0
                week_iso = r.get("week")
                year_val = None
                month_val = None
                weeknum_val = None
                phase_name = None
                if week_iso:
                    try:
                        wdate = _to_date(week_iso)
                        year_val = wdate.year
                        month_val = wdate.month
                        # Use ISO week number (equivalent to Excel WEEKNUM with return_type=21)
                        try:
                            weeknum_val = wdate.isocalendar().week  # type: ignore[attr-defined]
                        except Exception:
                            weeknum_val = wdate.isocalendar()[1]
                        # find matching phase window: start <= date <= end
                        for (s_dt, e_dt, nm) in phase_windows:
                            if s_dt <= wdate <= e_dt:
                                phase_name = nm
                                break
                    except Exception:
                        pass

                # Write values
                vals = [name, wc_name, mw, year_val, month_val, weeknum_val, phase_name]
                for c_idx, val in enumerate(vals, start=1):
                    cell = ws_list.cell(row=row_idx, column=c_idx)
                    cell.value = val
                    # Apply prototype formatting if available
                    if 1 <= c_idx <= len(proto_styles):
                        font, fill, border, alignment, number_format, protection = proto_styles[c_idx - 1]
                        if font is not None:
                            cell.font = font
                        if fill is not None:
                            cell.fill = fill
                        if border is not None:
                            cell.border = border
                        if alignment is not None:
                            cell.alignment = alignment
                        if number_format is not None:
                            cell.number_format = number_format
                        if protection is not None:
                            cell.protection = protection
                row_idx += 1
        except Exception:
            pass

        # Save as new file
        # Ensure directory exists
        os.makedirs(os.path.dirname(save_path) or ".", exist_ok=True)
        try:
            wb.save(save_path)
        except PermissionError:
            # Common on Windows if the file is open in Excel
            try:
                wb.close()
            except Exception:
                pass
            return {
                "success": False,
                "error": "The file is currently open and cannot be written. Please close the file and try again.",
            }
        except Exception as e:
            # Ensure the workbook is closed before bubbling up
            try:
                wb.close()
            except Exception:
                pass
            return {"success": False, "error": str(e)}
        # Explicitly close workbook to release file handle on success
        try:
            wb.close()
        except Exception:
            pass

        return {"success": True, "path": save_path}
    except Exception as e:
        return {"success": False, "error": str(e)}


def _sanitize_filename(s: str) -> str:
    import re
    return re.sub(r"[^\w\-_. ]", "_", s or "")


def suggest_csv_filename(payload: dict) -> str:
    """Suggest a default CSV filename using subproject name and week range."""
    records = (payload or {}).get("records") or []
    sub = (payload or {}).get("subproject") or {}
    sp_name = _sanitize_filename(sub.get("name") or f"Subproject_{sub.get('id')}")
    weeks = sorted({r.get("week") for r in records if r.get("week")})
    if weeks:
        return f"PMM_{sp_name}_{weeks[0]}_{weeks[-1]}.csv"
    return f"PMM_{sp_name}.csv"


def suggest_xlsx_filename(payload: dict) -> str:
    """Suggest a default XLSX filename using subproject name and week range."""
    records = (payload or {}).get("records") or []
    sub = (payload or {}).get("subproject") or {}
    sp_name = _sanitize_filename(sub.get("name") or f"Subproject_{sub.get('id')}")
    weeks = sorted({r.get("week") for r in records if r.get("week")})
    if weeks:
        return f"PMM_{sp_name}_{weeks[0]}_{weeks[-1]}.xlsx"
    return f"PMM_{sp_name}.xlsx"


def export_pmm_workloads_to_csv(payload: dict, save_path: str) -> Dict[str, Any]:
    """Export PMM workloads as a pivot CSV (rows=Work Category, columns=Week)."""
    try:
        records = (payload or {}).get("records") or []
        if not records:
            return {"success": False, "error": "No records"}

        # Compute weeks and matrix
        weeks = sorted({r.get("week") for r in records if r.get("week")})
        from collections import defaultdict
        matrix = defaultdict(lambda: {w: 0.0 for w in weeks})
        for r in records:
            week = r.get("week")
            if not week:
                continue
            wc = r.get("work_category") or {}
            name = wc.get("name") or "Unassigned"
            try:
                val = float(r.get("man_week") or 0)
            except Exception:
                val = 0.0
            matrix[name][week] = matrix[name].get(week, 0.0) + val

        category_names = sorted(matrix.keys(), key=lambda s: (s is None, str(s)))

        os.makedirs(os.path.dirname(save_path) or ".", exist_ok=True)
        try:
            with open(save_path, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                header = ["Work Category", *weeks]
                writer.writerow(header)
                for name in category_names:
                    row = [name]
                    row.extend([matrix[name].get(w, 0.0) for w in weeks])
                    writer.writerow(row)
        except PermissionError:
            return {
                "success": False,
                "error": "The file is currently open and cannot be written. Please close the file and try again.",
            }
        return {"success": True, "path": save_path}
    except Exception as e:
        return {"success": False, "error": str(e)}
